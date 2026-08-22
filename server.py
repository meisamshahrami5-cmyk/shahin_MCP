import os
import re
import time
import logging
from pathlib import Path
from datetime import datetime
from mcp.server import MCPServer
from openpyxl import Workbook, load_workbook


mcp = MCPServer(
    name="my-mcp",
    version="1.0.0",
)


# ─────────────────────────────────────────────
# تنظیمات اصلی
# ─────────────────────────────────────────────

BASE_DIR = Path(r"C:\Users\AVANG\Desktop\my-mcp").resolve()

# مسیرهای ممنوعه (حتی اگر داخل BASE_DIR هم نباشند، چک می‌شن)
FORBIDDEN_PATHS = [
    Path(r"C:\Windows"),
    Path(r"C:\Program Files"),
    Path(r"C:\Program Files (x86)"),
    Path(os.path.expanduser("~/.ssh")),
    Path(os.path.expanduser("~/.aws")),
    Path(os.path.expanduser("~/.config")),
]

# پسوندهای مجاز برای خواندن و نوشتن
ALLOWED_EXTENSIONS = {
    ".txt", ".py", ".json", ".md", ".html",
    ".css", ".js", ".csv", ".xml", ".yaml", ".yml",
}

# الگوهای خطرناک در محتوای فایل‌ها
DANGEROUS_PATTERNS = [
    r"rm\s+-rf\s+/",
    r"sudo\s+",
    r"chmod\s+777",
    r">\s+/dev/",
    r"mkfs",
    r"/etc/passwd",
    r"/etc/shadow",
    r"\.ssh/",
    r"\.aws/",
    r":()\{.*\|.*\&.*\}",  # fork bomb
]

# محدودیت نرخ عملیات
MAX_OPS_PER_MINUTE = 30
MAX_FILE_SIZE_READ  = 10 * 1024 * 1024  # 10 MB
MAX_FILE_SIZE_WRITE =  1 * 1024 * 1024  #  1 MB


# ─────────────────────────────────────────────
# سیستم لاگینگ
# ─────────────────────────────────────────────

log_dir = BASE_DIR / "logs"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / f"mcp_{datetime.now().strftime('%Y%m%d')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("my-mcp")


def security_alert(message: str, severity: str = "WARNING"):
    if severity == "CRITICAL":
        logger.critical(f"🚨 {message}")
    elif severity == "WARNING":
        logger.warning(f"⚠️ {message}")
    else:
        logger.info(f"ℹ️ {message}")


# ─────────────────────────────────────────────
# کلاس امنیتی
# ─────────────────────────────────────────────

class SecurityValidator:
    def __init__(self):
        self._ops_log: list[float] = []

    # --- مسیر امن ---
    def safe_path(self, file_path: str) -> Path:
        """مسیر را به مطلق تبدیل و داخل BASE_DIR بررسی می‌کند."""
        path = Path(file_path)
        if not path.is_absolute():
            path = BASE_DIR / path
        path = path.resolve()

        # داخل BASE_DIR باشد
        try:
            path.relative_to(BASE_DIR)
        except ValueError:
            raise PermissionError("دسترسی خارج از پوشه پروژه مجاز نیست.")

        # جزو مسیرهای ممنوعه نباشد
        for forbidden in FORBIDDEN_PATHS:
            if str(path).startswith(str(forbidden.resolve())):
                raise PermissionError(f"دسترسی به مسیر ممنوعه مجاز نیست: {forbidden}")

        return path

    # --- اسکن محتوا ---
    def scan_content(self, content: str) -> bool:
        """True اگر محتوا خطرناک باشد."""
        for pattern in DANGEROUS_PATTERNS:
            if re.search(pattern, content, re.IGNORECASE):
                return True
        return False

    # --- محدودیت نرخ ---
    def check_rate_limit(self) -> bool:
        now = time.time()
        self._ops_log = [t for t in self._ops_log if now - t < 60]
        if len(self._ops_log) >= MAX_OPS_PER_MINUTE:
            return False
        self._ops_log.append(now)
        return True

    # --- بررسی پسوند ---
    def check_extension(self, path: Path) -> bool:
        return path.suffix.lower() in ALLOWED_EXTENSIONS

    # --- بررسی حجم فایل ---
    def check_size_read(self, path: Path) -> bool:
        return path.stat().st_size <= MAX_FILE_SIZE_READ

    def check_size_write(self, content: str) -> bool:
        return len(content.encode("utf-8")) <= MAX_FILE_SIZE_WRITE


# نمونه واحد validator
validator = SecurityValidator()


# ─────────────────────────────────────────────
# ابزارها
# ─────────────────────────────────────────────

@mcp.tool()
def hello(name: str) -> str:
    """به کاربر سلام می‌کند."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات: بیش از ۳۰ عملیات در دقیقه."
    return f"سلام {name}! MCP با موفقیت کار می‌کند."


@mcp.tool()
def list_files(folder: str) -> str:
    """فایل‌ها و پوشه‌های موجود در یک مسیر را نمایش می‌دهد."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."
    try:
        path = validator.safe_path(folder)
    except PermissionError as e:
        security_alert(f"list_files blocked: {e}")
        return f"⛔ {e}"

    if not path.exists():
        return f"مسیر پیدا نشد: {folder}"
    if not path.is_dir():
        return f"این مسیر یک پوشه نیست: {folder}"

    items = []
    for item in sorted(path.iterdir()):
        prefix = "[پوشه]" if item.is_dir() else "[فایل]"
        items.append(f"{prefix} {item.name}")

    return "\n".join(items) if items else "این پوشه خالی است."


@mcp.tool()
def read_file(file_path: str) -> str:
    """محتوای فایل متنی یا کدنویسی را می‌خواند."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."
    try:
        path = validator.safe_path(file_path)
    except PermissionError as e:
        security_alert(f"read_file blocked: {e}")
        return f"⛔ {e}"

    if not path.exists():
        return f"فایل پیدا نشد: {file_path}"
    if not path.is_file():
        return f"این مسیر یک فایل نیست: {file_path}"
    if not validator.check_extension(path):
        return f"⛔ این نوع فایل مجاز نیست: {path.suffix}"
    if not validator.check_size_read(path):
        return f"⛔ حجم فایل بیش از حد مجاز است (max {MAX_FILE_SIZE_READ // (1024*1024)} MB)."

    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return "فایل با UTF-8 قابل خواندن نیست."
    except Exception as e:
        return f"خطا: {e}"


@mcp.tool()
def write_file(file_path: str, content: str) -> str:
    """یک فایل متنی را ایجاد یا ویرایش می‌کند."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."

    # اسکن محتوا
    if validator.scan_content(content):
        security_alert(f"write_file: محتوای خطرناک در {file_path}", "CRITICAL")
        return "⛔ محتوای خطرناک شناسایی شد. عملیات متوقف شد."

    # بررسی حجم
    if not validator.check_size_write(content):
        return f"⛔ حجم محتوا بیش از حد مجاز است (max {MAX_FILE_SIZE_WRITE // 1024} KB)."

    try:
        path = validator.safe_path(file_path)
    except PermissionError as e:
        security_alert(f"write_file blocked: {e}")
        return f"⛔ {e}"

    if not validator.check_extension(path):
        return f"⛔ این نوع فایل مجاز نیست: {path.suffix}"

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")
        security_alert(f"write_file OK: {path}", "INFO")
        return f"✅ فایل ذخیره شد: {path}"
    except Exception as e:
        return f"خطا هنگام ذخیره فایل: {e}"


# ─── Excel ────────────────────────────────────

@mcp.tool()
def create_excel(filename: str, sheet_name: str = "Sheet1") -> str:
    """یک فایل Excel جدید ایجاد می‌کند."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    try:
        path = validator.safe_path(filename)
    except PermissionError as e:
        security_alert(f"create_excel blocked: {e}")
        return f"⛔ {e}"

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(path)
        security_alert(f"create_excel OK: {path}", "INFO")
        return f"✅ فایل Excel ساخته شد: {path}"
    except Exception as e:
        return f"خطا هنگام ساخت Excel: {e}"


@mcp.tool()
def write_excel(filename: str, sheet_name: str, cell: str, value: str) -> str:
    """یک مقدار را داخل یک سلول Excel می‌نویسد."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."

    if validator.scan_content(value):
        security_alert(f"write_excel: محتوای خطرناک در {filename}", "CRITICAL")
        return "⛔ محتوای خطرناک شناسایی شد."

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    try:
        path = validator.safe_path(filename)
    except PermissionError as e:
        security_alert(f"write_excel blocked: {e}")
        return f"⛔ {e}"

    if not path.exists():
        return f"فایل Excel پیدا نشد: {path}"

    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        ws[cell] = value
        wb.save(path)
        security_alert(f"write_excel OK: {path} [{sheet_name}!{cell}]", "INFO")
        return f"✅ مقدار در سلول {cell} از Sheet «{sheet_name}» نوشته شد."
    except Exception as e:
        return f"خطا هنگام نوشتن Excel: {e}"


@mcp.tool()
def read_excel(filename: str, sheet_name: str) -> str:
    """اطلاعات یک Sheet از فایل Excel را می‌خواند."""
    if not validator.check_rate_limit():
        return "⛔ محدودیت نرخ عملیات."

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    try:
        path = validator.safe_path(filename)
    except PermissionError as e:
        security_alert(f"read_excel blocked: {e}")
        return f"⛔ {e}"

    if not path.exists():
        return f"فایل Excel پیدا نشد: {path}"
    if not validator.check_size_read(path):
        return f"⛔ حجم فایل بیش از حد مجاز است."

    try:
        wb = load_workbook(path, data_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Sheet پیدا نشد: {sheet_name}"
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            if any(values):
                rows.append(" | ".join(values))
        return "\n".join(rows) if rows else "این Sheet خالی است."
    except Exception as e:
        return f"خطا هنگام خواندن Excel: {e}"


if __name__ == "__main__":
    mcp.run()
